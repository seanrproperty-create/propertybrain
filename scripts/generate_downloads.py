#!/usr/bin/env python3
"""
generate_propertybrain_downloads.py — Brief 405 Task 4
Generates 4 Excel spreadsheet downloads for propertybrain.uk

Run:  python3 scripts/generate_downloads.py

Output files written to:  assets/downloads/
"""
import os, sys

try:
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, LineChart, Reference
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl --quiet")
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, LineChart, Reference

OUT_DIR = os.path.join(os.path.dirname(__file__), '..', 'assets', 'downloads')
os.makedirs(OUT_DIR, exist_ok=True)

# ── Colour helpers ────────────────────────────────────────────────────────────
DARK_BLUE = 'FF1a365d'
AMBER     = 'FFd97706'
LIGHT_BLUE= 'FFe8f0fe'
GREEN     = 'FF16a34a'
RED       = 'FFdc2626'
WHITE     = 'FFFFFFFF'
GREY_BG   = 'FFf8f9fa'

def hdr_fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def hdr_font(bold=True, color='FFFFFFFF', size=11):
    return Font(bold=bold, color=color, name='Calibri', size=size)

def body_font(bold=False, color='FF333333', size=10):
    return Font(bold=bold, color=color, name='Calibri', size=size)

def thin_border():
    s = Side(border_style='thin', color='FFCCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal='center', vertical='center')

def set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

def add_brand_header(ws, title, subtitle='PropertyBrain.uk — Free UK Property Tools'):
    ws.merge_cells('A1:H1')
    c = ws['A1']
    c.value = title
    c.font = Font(bold=True, color=WHITE, size=14, name='Calibri')
    c.fill = hdr_fill(DARK_BLUE)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 32

    ws.merge_cells('A2:H2')
    c = ws['A2']
    c.value = subtitle
    c.font = Font(bold=False, color=DARK_BLUE[2:], size=10, name='Calibri', italic=True)
    c.fill = hdr_fill('FFe8f0fe')
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 20


# ── Download 1: BTL Cashflow Tracker ─────────────────────────────────────────
def build_btl_tracker():
    wb = openpyxl.Workbook()

    # ── Sheet 1: Monthly Tracker ──
    ws = wb.active
    ws.title = 'Monthly Tracker'
    add_brand_header(ws, 'Buy-to-Let Cashflow Tracker 2026', 'PropertyBrain.uk — Free UK Property Tools')

    MONTHS = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
              'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

    headers = ['', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
               'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec', 'TOTAL']
    row4 = ws.append
    ws.append([None])  # blank row 3

    # Header row
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=col_idx, value=h)
        if col_idx > 1:
            c.fill = hdr_fill(DARK_BLUE)
            c.font = hdr_font(color='FFFFFFFF', size=10)
            c.alignment = center()
        else:
            c.fill = hdr_fill(DARK_BLUE)
            c.font = hdr_font(color='FFFFFFFF', size=10)
        c.border = thin_border()
    ws.row_dimensions[4].height = 24

    # Data rows
    income_rows = [
        ('INCOME', None),
        ('Rent Received (£)', 900),
        ('Other Income (£)', 0),
        ('Total Income', None),  # formula row
    ]
    expense_rows = [
        ('EXPENSES', None),
        ('Void / Empty Days (£)', 0),
        ('Mortgage Payment (£)', 600),
        ('Management Fee (£)', 90),
        ('Maintenance & Repairs (£)', 83),
        ('Insurance (£)', 67),
        ('Ground Rent (£)', 0),
        ('Service Charge (£)', 0),
        ('Council Tax (empty periods) (£)', 0),
        ('Accountancy & Legal (£)', 42),
        ('Other Expenses (£)', 0),
        ('Total Expenses', None),
    ]
    cashflow_rows = [
        ('NET CASHFLOW', None),
    ]

    row = 5
    income_start = row
    for label, default in income_rows:
        is_total = label.startswith('Total')
        is_header = label.isupper() and default is None and not is_total
        for col_idx in range(1, 15):
            c = ws.cell(row=row, column=col_idx)
            c.border = thin_border()
            if col_idx == 1:
                c.value = label
                if is_header:
                    c.fill = hdr_fill(AMBER)
                    c.font = hdr_font(color='FFFFFFFF', size=10)
                elif is_total:
                    c.fill = hdr_fill(LIGHT_BLUE)
                    c.font = body_font(bold=True)
                else:
                    c.font = body_font()
            elif col_idx <= 13 and not is_header:
                if is_total:
                    # Sum of income lines in this month
                    prev = row - 1
                    start = income_start + 1
                    c.value = f'={get_column_letter(col_idx)}{start}+{get_column_letter(col_idx)}{start+1}'
                    c.fill = hdr_fill(LIGHT_BLUE)
                    c.font = body_font(bold=True)
                    c.number_format = '£#,##0.00'
                elif default is not None:
                    c.value = default
                    c.number_format = '£#,##0.00'
                    c.fill = hdr_fill(GREY_BG)
                c.alignment = center()
            elif col_idx == 14 and not is_header:
                if is_total or (default is not None):
                    c.value = f'=SUM(B{row}:M{row})'
                    c.number_format = '£#,##0.00'
                    c.font = body_font(bold=True)
                    c.fill = hdr_fill(LIGHT_BLUE) if is_total else hdr_fill(GREY_BG)
                c.alignment = center()
        row += 1

    total_income_row = row - 1

    exp_start = row
    for label, default in expense_rows:
        is_total = label.startswith('Total')
        is_header = label.isupper() and default is None and not is_total
        for col_idx in range(1, 15):
            c = ws.cell(row=row, column=col_idx)
            c.border = thin_border()
            if col_idx == 1:
                c.value = label
                if is_header:
                    c.fill = hdr_fill(AMBER)
                    c.font = hdr_font(color='FFFFFFFF', size=10)
                elif is_total:
                    c.fill = hdr_fill(LIGHT_BLUE)
                    c.font = body_font(bold=True)
                else:
                    c.font = body_font()
            elif col_idx <= 13 and not is_header:
                if is_total:
                    start = exp_start + 1
                    end = row - 1
                    c.value = f'=SUM({get_column_letter(col_idx)}{start}:{get_column_letter(col_idx)}{end})'
                    c.fill = hdr_fill(LIGHT_BLUE)
                    c.font = body_font(bold=True)
                    c.number_format = '£#,##0.00'
                elif default is not None:
                    c.value = default
                    c.number_format = '£#,##0.00'
                    c.fill = hdr_fill(GREY_BG)
                c.alignment = center()
            elif col_idx == 14 and not is_header:
                if is_total or (default is not None):
                    c.value = f'=SUM(B{row}:M{row})'
                    c.number_format = '£#,##0.00'
                    c.font = body_font(bold=True)
                    c.fill = hdr_fill(LIGHT_BLUE) if is_total else hdr_fill(GREY_BG)
                c.alignment = center()
        row += 1

    total_exp_row = row - 1

    # Net cashflow row
    for col_idx in range(1, 15):
        c = ws.cell(row=row, column=col_idx)
        c.border = thin_border()
        if col_idx == 1:
            c.value = 'NET CASHFLOW (£)'
            c.font = body_font(bold=True, size=11)
            c.fill = hdr_fill(DARK_BLUE)
            c.font = Font(bold=True, color='FFFFFFFF', size=11, name='Calibri')
        elif col_idx <= 13:
            c.value = f'={get_column_letter(col_idx)}{total_income_row}-{get_column_letter(col_idx)}{total_exp_row}'
            c.number_format = '£#,##0.00'
            c.font = body_font(bold=True)
            c.fill = hdr_fill(DARK_BLUE)
            c.font = Font(bold=True, color='FFFFFFFF', size=11, name='Calibri')
            # Conditional formatting can't be added simply here; leave plain
        elif col_idx == 14:
            c.value = f'=SUM(B{row}:M{row})'
            c.number_format = '£#,##0.00'
            c.fill = hdr_fill(DARK_BLUE)
            c.font = Font(bold=True, color='FFFFFFFF', size=11, name='Calibri')
        c.alignment = center()
    ws.row_dimensions[row].height = 26

    # Column widths
    ws.column_dimensions['A'].width = 28
    for col in 'BCDEFGHIJKLMN':
        ws.column_dimensions[col].width = 10

    ws.freeze_panes = 'B5'

    # ── Sheet 2: Annual Summary ──
    ws2 = wb.create_sheet('Annual Summary')
    add_brand_header(ws2, 'Annual Summary & Yield Calculator')
    ws2.append([None])
    inputs = [
        ('PROPERTY DETAILS', None, None),
        ('Purchase Price', 200000, '£#,##0'),
        ('Monthly Rent (target)', 900, '£#,##0'),
        ('Annual Costs (estimated)', 3600, '£#,##0'),
        ('', None, None),
        ('ANNUAL RESULTS', None, None),
        ('Annual Gross Rent', '=B4*12', '£#,##0'),
        ('Annual Costs', '=B5', '£#,##0'),
        ('Net Annual Income', '=B10-B11', '£#,##0'),
        ('', None, None),
        ('YIELD', None, None),
        ('Gross Yield', '=B10/B3', '0.00%'),
        ('Net Yield', '=B12/B3', '0.00%'),
    ]
    for row_data in inputs:
        label, val, fmt = row_data
        ws2.append([label, val])
        r = ws2.max_row
        cell_a = ws2[f'A{r}']
        cell_b = ws2[f'B{r}']
        if label and label == label.upper() and val is None and label != '':
            cell_a.fill = hdr_fill(DARK_BLUE)
            cell_a.font = Font(bold=True, color='FFFFFFFF', size=10, name='Calibri')
            ws2.merge_cells(f'A{r}:C{r}')
        else:
            cell_a.font = body_font()
            if val is not None or (isinstance(val, str) and val.startswith('=')):
                cell_b.value = val
                if fmt:
                    cell_b.number_format = fmt
                cell_b.font = body_font(bold=(fmt and 'ANNUAL' in str(fmt)))
        cell_a.border = thin_border()
        cell_b.border = thin_border()

    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 15

    out = os.path.join(OUT_DIR, 'btl-cashflow-tracker.xlsx')
    wb.save(out)
    print(f'  Created: {os.path.basename(out)} ({os.path.getsize(out)//1024}KB)')


# ── Download 2: Property Deal Analyser ────────────────────────────────────────
def build_deal_analyser():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Deal Analyser'
    add_brand_header(ws, 'Property Deal Analyser', 'PropertyBrain.uk — Enter details below to auto-calculate all metrics')

    ws.append([None])

    sections = [
        # (label, input_cell_value, format, is_section_header)
        ('PURCHASE DETAILS', None, None, True),
        ('Purchase Price (£)', 200000, '£#,##0', False),
        ('Deposit (%)', 25, '0%', False),
        ('Mortgage Rate (%)', 5.0, '0.00%', False),
        ('Mortgage Term (years)', 25, '0', False),
        ('Stamp Duty (£)', '=IF(B5<=250000,0,IF(B5<=925000,(B5-250000)*0.05,33750+(B5-925000)*0.1))*1.05+(B5*0.05)',
         '£#,##0', False),
        ('Legal & Survey Costs (£)', 2500, '£#,##0', False),
        ('', None, None, False),
        ('RENTAL INCOME', None, None, True),
        ('Monthly Rent (£)', 900, '£#,##0', False),
        ('Void Allowance (%)', 8, '0%', False),
        ('Net Monthly Rent', '=B11*(1-B12)', '£#,##0', False),
        ('', None, None, False),
        ('ANNUAL COSTS', None, None, True),
        ('Management Fee (% of rent)', 10, '0%', False),
        ('Maintenance (£/year)', 1200, '£#,##0', False),
        ('Insurance (£/year)', 800, '£#,##0', False),
        ('Ground Rent (£/year)', 0, '£#,##0', False),
        ('Service Charge (£/year)', 0, '£#,##0', False),
        ('Total Annual Costs', '=(B11*12*B16)+(B17+B18+B19+B20)', '£#,##0', False),
        ('', None, None, False),
        ('RESULTS', None, None, True),
        ('Gross Yield', '=B11*12/B5', '0.00%', False),
        ('Net Yield', '=(B11*12-B22)/B5', '0.00%', False),
        ('Loan Amount', '=B5*(1-B6)', '£#,##0', False),
        ('Monthly Mortgage (IO)', '=B25*B7/12', '£#,##0', False),
        ('Monthly Mortgage (Repayment)', '=IFERROR(B25*(B7/12)/(1-(1+B7/12)^(-B8*12)),0)', '£#,##0', False),
        ('Monthly Net Cashflow (IO)', '=B13-B27/12-B26', '£#,##0', False),
        ('Monthly Net Cashflow (Repayment)', '=B13-B27/12-B27', '£#,##0', False),
        ('Cash Invested (deposit+costs)', '=B5*B6+B7+B8', '£#,##0', False),
        ('Cash-on-Cash Return (IO)', '=IF(B31>0,(B13-B27/12-B26)*12/B31,0)', '0.00%', False),
        ('Break-even Rent (IO)', '=(B26+B27/12)/(1-B12)', '£#,##0', False),
    ]

    for label, val, fmt, is_hdr in sections:
        row = ws.max_row + 1
        c_a = ws.cell(row=row, column=1, value=label)
        c_b = ws.cell(row=row, column=2, value=val)
        c_c = ws.cell(row=row, column=3)

        if is_hdr:
            c_a.fill = hdr_fill(DARK_BLUE)
            c_a.font = Font(bold=True, color='FFFFFFFF', size=10, name='Calibri')
            ws.merge_cells(f'A{row}:D{row}')
        elif label:
            c_a.font = body_font()
            if val is not None:
                c_b.fill = hdr_fill('FFfefce8') if not str(val).startswith('=') else hdr_fill(LIGHT_BLUE)
                c_b.font = body_font(bold=str(val).startswith('='))
                if fmt:
                    c_b.number_format = fmt

        for c in [c_a, c_b]:
            c.border = thin_border()

    # Add note column
    notes = {
        5: 'Enter purchase price',
        6: '25% typical for BTL',
        7: 'Annual rate e.g. 0.05 = 5%',
        11: 'Target monthly rent',
        12: 'Empty weeks per year / 52',
        16: '10-15% typical',
    }

    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 28

    # Add instructions note (row 3 is first unmerged row after 2-row header)
    c = ws.cell(row=3, column=3, value='← Yellow cells = inputs. Blue = calculated.')
    c.font = Font(italic=True, color='FF666666', size=9, name='Calibri')

    out = os.path.join(OUT_DIR, 'property-deal-analyser.xlsx')
    wb.save(out)
    print(f'  Created: {os.path.basename(out)} ({os.path.getsize(out)//1024}KB)')


# ── Download 3: Portfolio Tracker ────────────────────────────────────────────
def build_portfolio_tracker():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Portfolio'
    add_brand_header(ws, 'Property Portfolio Tracker 2026', 'PropertyBrain.uk — Track all your properties in one place')

    ws.append([None])

    cols = [
        ('Property Address', 40),
        ('Purchase Price', 14),
        ('Current Value', 14),
        ('Mortgage Balance', 14),
        ('Equity', 14),
        ('Monthly Rent', 13),
        ('Monthly Costs', 13),
        ('Net Cashflow', 13),
        ('Gross Yield', 12),
        ('Net Yield', 12),
        ('Notes', 22),
    ]

    # Header
    for col_idx, (header, width) in enumerate(cols, start=1):
        c = ws.cell(row=4, column=col_idx, value=header)
        c.fill = hdr_fill(DARK_BLUE)
        c.font = Font(bold=True, color='FFFFFFFF', size=10, name='Calibri')
        c.alignment = center()
        c.border = thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[4].height = 24

    # Sample data rows (5 properties)
    samples = [
        ['23 High Street, Sheffield', 150000, 165000, 112500, None, 750, 300, None, None, None, 'BTL - terraced'],
        ['14 Park Ave, Manchester', 200000, 215000, 150000, None, 900, 380, None, None, None, 'BTL - semi'],
        ['', None, None, None, None, None, None, None, None, None, ''],
        ['', None, None, None, None, None, None, None, None, None, ''],
        ['', None, None, None, None, None, None, None, None, None, ''],
    ]

    for i, sample in enumerate(samples, start=5):
        for col_idx, val in enumerate(sample, start=1):
            c = ws.cell(row=i, column=col_idx, value=val)
            c.border = thin_border()
            c.fill = hdr_fill(GREY_BG) if i % 2 == 0 else hdr_fill(WHITE)
            c.font = body_font()
            if col_idx == 5 and i <= 6:
                c.value = f'=C{i}-D{i}'
                c.number_format = '£#,##0'
            elif col_idx == 8 and i <= 6:
                c.value = f'=F{i}-G{i}'
                c.number_format = '£#,##0'
            elif col_idx == 9 and i <= 6:
                c.value = f'=IF(B{i}>0,F{i}*12/B{i},"")'
                c.number_format = '0.00%'
            elif col_idx == 10 and i <= 6:
                c.value = f'=IF(B{i}>0,H{i}*12/B{i},"")'
                c.number_format = '0.00%'
            elif col_idx in (2, 3, 4, 5, 6, 7, 8) and val is not None:
                c.number_format = '£#,##0'
            c.alignment = Alignment(horizontal='center' if col_idx > 1 else 'left', vertical='center')

    # Summary row
    sum_row = 10
    ws.cell(row=sum_row, column=1, value='PORTFOLIO TOTAL').font = Font(bold=True, size=11, name='Calibri', color=WHITE)
    ws.cell(row=sum_row, column=1).fill = hdr_fill(DARK_BLUE)
    ws.cell(row=sum_row, column=1).border = thin_border()
    for col_idx in range(2, 12):
        c = ws.cell(row=sum_row, column=col_idx)
        c.fill = hdr_fill(DARK_BLUE)
        c.font = Font(bold=True, color='FFFFFFFF', size=10, name='Calibri')
        c.border = thin_border()
        c.alignment = center()
        if col_idx in (2, 3, 4, 5, 6, 7, 8):
            c.value = f'=SUM({get_column_letter(col_idx)}5:{get_column_letter(col_idx)}9)'
            c.number_format = '£#,##0'
        elif col_idx == 9:
            c.value = f'=IFERROR(F{sum_row}*12/B{sum_row},"")'
            c.number_format = '0.00%'

    ws.freeze_panes = 'A5'

    out = os.path.join(OUT_DIR, 'property-portfolio-tracker.xlsx')
    wb.save(out)
    print(f'  Created: {os.path.basename(out)} ({os.path.getsize(out)//1024}KB)')


# ── Download 4: SA Income Calculator ────────────────────────────────────────
def build_sa_calculator():
    wb = openpyxl.Workbook()

    # Sheet 1: SA Calculator
    ws = wb.active
    ws.title = 'SA Calculator'
    add_brand_header(ws, 'Serviced Accommodation Income Calculator', 'PropertyBrain.uk — Compare SA vs BTL returns')

    ws.append([None])

    inputs = [
        ('PROPERTY & RATES', None, None, True),
        ('Nightly Rate (£)', 85, '£#,##0', False),
        ('Platform fee (%)', 3, '0%', False),
        ('Target Occupancy (%)', 65, '0%', False),
        ('Nights per Month (avg)', 30, '0', False),
        ('', None, None, False),
        ('MONTHLY COSTS', None, None, True),
        ('Cleaning per stay (£)', 60, '£#,##0', False),
        ('Avg stays per month', 8, '0', False),
        ('Management fee (% of revenue)', 12, '0%', False),
        ('Utilities (£/mo)', 150, '£#,##0', False),
        ('Insurance (£/mo)', 80, '£#,##0', False),
        ('Mortgage (£/mo)', 600, '£#,##0', False),
        ('', None, None, False),
        ('BTL COMPARISON', None, None, True),
        ('Equivalent BTL monthly rent (£)', 750, '£#,##0', False),
        ('BTL management fee (%)', 10, '0%', False),
        ('BTL other monthly costs (£)', 200, '£#,##0', False),
        ('', None, None, False),
        ('SA RESULTS', None, None, True),
        ('Gross Monthly Revenue', '=B5*B6*B7', '£#,##0', False),
        ('Less: Platform Fee', '=B22*(B6)', '£#,##0', False),
        ('Net Revenue', '=B22-B23', '£#,##0', False),
        ('Cleaning Costs', '=B10*B11', '£#,##0', False),
        ('Management Fee', '=B22*B12', '£#,##0', False),
        ('Utilities', '=B13', '£#,##0', False),
        ('Insurance', '=B14', '£#,##0', False),
        ('Mortgage', '=B15', '£#,##0', False),
        ('Total Monthly Costs', '=SUM(B25:B29)', '£#,##0', False),
        ('NET MONTHLY PROFIT (SA)', '=B24-B30', '£#,##0', False),
        ('Annual SA Income', '=B31*12', '£#,##0', False),
        ('', None, None, False),
        ('BTL RESULTS', None, None, True),
        ('BTL Net Monthly Income', '=B17*(1-B18)-B19', '£#,##0', False),
        ('Annual BTL Income', '=B35*12', '£#,##0', False),
        ('', None, None, False),
        ('SA UPLIFT vs BTL', None, None, True),
        ('Extra monthly income (SA vs BTL)', '=B31-B35', '£#,##0', False),
        ('Extra annual income (SA vs BTL)', '=B39*12', '£#,##0', False),
    ]

    for label, val, fmt, is_hdr in inputs:
        row = ws.max_row + 1
        c_a = ws.cell(row=row, column=1, value=label)
        c_b = ws.cell(row=row, column=2, value=val)
        if is_hdr:
            c_a.fill = hdr_fill(DARK_BLUE)
            c_a.font = Font(bold=True, color='FFFFFFFF', size=10, name='Calibri')
            ws.merge_cells(f'A{row}:D{row}')
        elif label:
            c_a.font = body_font()
            if val is not None:
                is_calc = isinstance(val, str) and val.startswith('=')
                c_b.fill = hdr_fill('FFfefce8') if not is_calc else hdr_fill(LIGHT_BLUE)
                c_b.font = body_font(bold=is_calc)
                if fmt:
                    c_b.number_format = fmt
        for c in [c_a, c_b]:
            c.border = thin_border()

    # Sheet 2: Seasonality
    ws2 = wb.create_sheet('Seasonality')
    add_brand_header(ws2, 'SA Seasonality — 12-Month Projection')
    ws2.append([None])

    months = ['January', 'February', 'March', 'April', 'May', 'June',
              'July', 'August', 'September', 'October', 'November', 'December']
    default_occ = [55, 55, 65, 70, 72, 78, 82, 82, 70, 65, 55, 60]

    hdr_cols = ['Month', 'Occupancy %', 'Nights Occupied', 'Gross Revenue', 'Platform Fee', 'Net Revenue',
                'Costs', 'Net Profit']
    for ci, h in enumerate(hdr_cols, start=1):
        c = ws2.cell(row=4, column=ci, value=h)
        c.fill = hdr_fill(DARK_BLUE)
        c.font = Font(bold=True, color='FFFFFFFF', size=10, name='Calibri')
        c.alignment = center()
        c.border = thin_border()

    # Reference cells from Sheet 1 (using hardcoded defaults here for standalone)
    nightly = 85
    platform = 0.03
    costs_pm = 600 + 80 + 150  # mortgage + insurance + utilities

    for i, (month, occ) in enumerate(zip(months, default_occ), start=5):
        nights = round(30 * occ / 100)
        gross = nights * nightly
        fee = gross * platform
        net_rev = gross - fee
        cleaning = 8 * 60  # fixed example
        mgmt = gross * 0.12
        total_costs = cleaning + mgmt + costs_pm
        net_profit = net_rev - total_costs

        row_data = [month, occ/100, nights, gross, fee, net_rev, total_costs, net_profit]
        fmts = ['', '0%', '0', '£#,##0', '£#,##0', '£#,##0', '£#,##0', '£#,##0']
        for ci, (val, fmt) in enumerate(zip(row_data, fmts), start=1):
            c = ws2.cell(row=i, column=ci, value=val)
            if fmt:
                c.number_format = fmt
            c.font = body_font()
            c.fill = hdr_fill(GREY_BG) if i % 2 == 0 else hdr_fill(WHITE)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = thin_border()

    # Totals
    tot_row = 17
    c = ws2.cell(row=tot_row, column=1, value='ANNUAL TOTAL')
    c.fill = hdr_fill(DARK_BLUE)
    c.font = Font(bold=True, color='FFFFFFFF', size=10, name='Calibri')
    c.border = thin_border()
    for ci in range(2, 9):
        c = ws2.cell(row=tot_row, column=ci)
        if ci >= 3:
            c.value = f'=SUM({get_column_letter(ci)}5:{get_column_letter(ci)}16)'
            c.number_format = '£#,##0' if ci >= 4 else '0'
            c.font = Font(bold=True, color='FFFFFFFF', size=10, name='Calibri')
        c.fill = hdr_fill(DARK_BLUE)
        c.border = thin_border()
        c.alignment = center()

    ws2.column_dimensions['A'].width = 15
    for col in 'BCDEFGH':
        ws2.column_dimensions[col].width = 14

    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 16

    out = os.path.join(OUT_DIR, 'sa-income-calculator.xlsx')
    wb.save(out)
    print(f'  Created: {os.path.basename(out)} ({os.path.getsize(out)//1024}KB)')


if __name__ == '__main__':
    print('Generating PropertyBrain.uk Excel downloads...')
    build_btl_tracker()
    build_deal_analyser()
    build_portfolio_tracker()
    build_sa_calculator()
    print('Done. All 4 files created in assets/downloads/')
